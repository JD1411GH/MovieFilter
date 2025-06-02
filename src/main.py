import json
import os
import dataset
import tkinter as tk
import openpyxl
import random
import subprocess
import threading

curdir = os.path.dirname(__file__)
rootdir = os.path.dirname(curdir)

filtered_movies = []
tkroot = None
filter_row_frame = None
movie_count_label = None

current_filters = {
    'studio': 'All',
    'category': 'All',
    'actor_rating': 'All',
    'actor': 'All',
    'movie_rating': 'All'
}

with open(os.path.join(rootdir, 'config.json')) as f:
    config = json.load(f)


def convert_xls_to_sqlite(table):
    table.delete()  
    workbook = openpyxl.load_workbook(filename='LockerDB.xlsx')
    sheet = workbook.active
    headers = {header: idx for idx, header in enumerate(
        next(sheet.iter_rows(values_only=True)), 1)}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert({
            'rel_path': row[headers['rel_path'] - 1],
            'movie_rating': row[headers['movie_rating'] - 1],
            'actor_rating': row[headers['actor_rating'] - 1],
            'actor': row[headers['actor'] - 1],
            'category': row[headers['category'] - 1],
            'studio': row[headers['studio'] - 1],
            'last_played': "",
            'playcount': row[headers['playcount'] - 1]
        })
    for row in table.all():
        updates = {}
        for col in ['movie_rating', 'actor_rating']:
            val = row.get(col)
            if val is not None:
                try:
                    updates[col] = int(float(val))
                except (ValueError, TypeError):
                    continue
        if updates:
            table.update(dict(id=row['id'], **updates), ['id'])


def create_action_row():
    global tkroot

    # Frame to hold the actor search input
    frame = tk.Frame(tkroot)
    frame.pack(pady=10)

    # Actor search label and entry
    actor_search_label = tk.Label(frame, text="Search Actor:")
    actor_search_label.pack(side=tk.LEFT)

    actor_search_var = tk.StringVar()
    actor_search_entry = tk.Entry(frame, textvariable=actor_search_var)
    actor_search_entry.pack(side=tk.LEFT, padx=5)

    # Search button
    search_button = tk.Button(frame, text="Search")
    search_button.pack(side=tk.LEFT, padx=5)

    # Buttons frame
    buttons_frame = tk.Frame(frame)
    buttons_frame.pack(side=tk.LEFT, padx=5)

    play_selected_button = tk.Button(
        buttons_frame, text="Play random", command=play_random)
    play_selected_button.pack(side=tk.LEFT, padx=2)

    # reset button
    reset_button = tk.Button(
        buttons_frame, text="Reset filters", command=reset_filters)
    reset_button.pack(side=tk.LEFT, padx=2)


def create_filter_row():
    global filtered_movies, tkroot, filter_row_frame, current_filters

    # Destroy previous filter row if it exists
    if filter_row_frame is not None:
        filter_row_frame.destroy()

    # Main frame to hold all filters
    filter_row_frame = tk.Frame(tkroot)
    if len(tkroot.winfo_children()) > 1:
        filter_row_frame.pack(pady=10, before=tkroot.winfo_children()[0])
    else:
        filter_row_frame.pack(pady=10)

    # Studio filter
    studio_frame = tk.Frame(filter_row_frame)
    studio_frame.pack(side=tk.LEFT, padx=10)
    studio_label = tk.Label(studio_frame, text="studio")
    studio_label.pack()
    studios = ["All"] + sorted({row['studio']
                                for row in filtered_movies if 'studio' in row and row['studio']})
    studio_var = tk.StringVar()
    studio_var.set(current_filters['studio'])
    studio_dropdown = tk.OptionMenu(
        studio_frame, studio_var, *studios, command=lambda value: on_filter_change('studio', value))
    studio_dropdown.pack()
    set_current_filters()
    create_movie_count()

    # Category filter
    category_frame = tk.Frame(filter_row_frame)
    category_frame.pack(side=tk.LEFT, padx=10)
    category_label = tk.Label(category_frame, text="category")
    category_label.pack()
    categories = ["All"] + sorted({row['category']
                                   for row in filtered_movies if 'category' in row and row['category']})
    category_var = tk.StringVar()
    category_var.set(categories[0])
    category_dropdown = tk.OptionMenu(
        category_frame, category_var, *categories, command=lambda value: on_filter_change('category', value))
    category_dropdown.pack()

    # actor_rating rating filter
    # Subframe for actor_rating filter
    actor_rating_frame = tk.Frame(filter_row_frame)
    actor_rating_frame.pack(side=tk.LEFT, padx=10)
    actor_rating_label = tk.Label(actor_rating_frame, text="actor_rating")
    actor_rating_label.pack()  # Default is TOP (vertical alignment)
    actor_ratings = ["All"] + sorted({row['actor_rating']
                                      for row in filtered_movies if 'actor_rating' in row and row['actor_rating']})
    actor_rating_var = tk.StringVar()
    actor_rating_var.set(actor_ratings[0])
    actor_rating_dropdown = tk.OptionMenu(
        actor_rating_frame, actor_rating_var, *actor_ratings, command=lambda value: on_filter_change('actor_rating', value))
    actor_rating_dropdown.pack()

    # actor filter
    actor_frame = tk.Frame(filter_row_frame)
    actor_frame.pack(side=tk.LEFT, padx=10)
    actor_label = tk.Label(actor_frame, text="actor")
    actor_label.pack()
    actors = ["All"] + sorted({row['actor']
                               for row in filtered_movies if 'actor' in row and row['actor']})
    actor_var = tk.StringVar()
    actor_var.set(actors[0])
    actor_dropdown = tk.OptionMenu(
        actor_frame, actor_var, *actors, command=lambda value: on_filter_change('actor', value))
    actor_dropdown.pack()

    # movie_rating filter
    movie_rating_frame = tk.Frame(filter_row_frame)
    movie_rating_frame.pack(side=tk.LEFT, padx=10)
    movie_rating_label = tk.Label(movie_rating_frame, text="movie_rating")
    movie_rating_label.pack()
    movie_ratings = ["All"] + sorted({row['movie_rating']
                                      for row in filtered_movies if 'movie_rating' in row and row['movie_rating']})
    movie_rating_var = tk.StringVar()
    movie_rating_var.set(movie_ratings[0])
    movie_rating_dropdown = tk.OptionMenu(
        movie_rating_frame, movie_rating_var, *movie_ratings, command=lambda value: on_filter_change('movie_rating', value))
    movie_rating_dropdown.pack()


def fetch_all_movies():
    global filtered_movies
    db = dataset.connect(f'sqlite:///{config["dbfile"]}')
    table = db['movies']
    # convert_xls_to_sqlite(table)
    filtered_movies = list(table.all())


def create_movie_count():
    global tkroot, movie_count_label
    if movie_count_label is not None:
        movie_count_label.destroy()
    movie_count_label = tk.Label(
        tkroot, text=f"Movie count: {len(filtered_movies)}")
    movie_count_label.pack(pady=10)


def on_filter_change(filter_type, value):
    global current_filters
    current_filters[filter_type] = value
    set_current_filters()
    create_movie_count()


def play_random():
    global filtered_movies
    if len(filtered_movies) == 0:
        print("No movies found")
        return

    def run_player():
        random.shuffle(filtered_movies)
        cmd = f"{config['movie_player']} {config['moviedir']}\\{filtered_movies[0]['rel_path']}"
        process = subprocess.Popen(cmd, shell=True)
        process.wait()
        print("movie is over")

    threading.Thread(target=run_player, daemon=True).start()


def reset_filters():
    global studio_var, current_filters
    fetch_all_movies()
    current_filters.clear()
    create_filter_row()


def set_current_filters():
    global filtered_movies
    fetch_all_movies()
    for key, value in current_filters.items():
        if value != "All":
            filtered_movies = [
                movie for movie in filtered_movies if movie.get(key) == value]
    print(f"number of filtered movies: {len(filtered_movies)}")


def main():
    global tkroot
    tkroot = tk.Tk()
    tkroot.title("Movie Filter")
    fetch_all_movies()
    create_filter_row()
    create_action_row()
    create_movie_count()
    tkroot.mainloop()


if __name__ == "__main__":
    main()
